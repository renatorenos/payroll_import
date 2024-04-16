from openpyxl import load_workbook
from loguru import logger

def parserFile_dict(file : str) -> dict:
    """
        Função que extrai os valores do xlsx .
    Args:
        param1: conteúdo do arquivo xlsx.
    Returns:
        Retorna um dicionários contendo {matricula : valor}.
    """
    wb = load_workbook(filename= file, data_only=True)
    ws = wb['Planilha1']

    matriculationCol = 'A'
    valueCol = 'B'
    logger.debug("Linhas da planilha[" + matriculationCol + "] = " + str(len(ws[matriculationCol])))   
    logger.debug("Linhas da planilha[" + valueCol + "] = " + str(len(ws[matriculationCol])))

    dic = {}
    for i in range(1, len(ws[matriculationCol]) + 1):
        key = ws[matriculationCol + str(i)].value
        if key is None:
            error = str(matriculationCol) + "[" + str(i) +"]" + " = Matricula em branco"
            logger.error(error)
            raise Exception("Ocorreu um erro => Matricula em branco")
        value = str(ws[valueCol + str(i)].value).replace(" ", "").replace(".","").replace(",","")
        
        logger.debug(f'[{i}] = {str(key).replace(" ", "")} : {value}')

        dic.update({ str(key).replace(" ", "") : value })
    
    return dic

def parserFile_list(file : str) -> list:
    """
        Função que extrai os valores do .xlsx.
    Args:
        param1: conteúdo do arquivo xlsx.
    Returns:
        Retorna uma lista contendo as tuplas (matricula, valor) .
    """
    wb = load_workbook(filename= file, data_only=True)
    ws = wb['Planilha1']

    matriculationCol = 'A'
    valueCol = 'B'
    logger.debug("Linhas da planilha[" + matriculationCol + "] = " + str(len(ws[matriculationCol])))   
    logger.debug("Linhas da planilha[" + valueCol + "] = " + str(len(ws[matriculationCol])))

    records = []
    for i in range(1, len(ws[matriculationCol]) + 1):
        key = ws[matriculationCol + str(i)].value
        if key is None:
            error = str(matriculationCol) + "[" + str(i) +"]" + " = Matricula em branco"
            logger.error(error)
            raise Exception("Ocorreu um erro => Matricula em branco")
        value = str(ws[valueCol + str(i)].value).replace(" ", "").replace(".","").replace(",","")
        k = str(key).replace(" ", "")

        logger.debug(f'[{i}] = {k} : {value}')
        records.append((k, value))
    
    return records